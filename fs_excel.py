# -*- coding: utf-8 -*-
"""
공시원문 → 재무제표 엑셀
================================================

내려받은 공시원문(zip) 안에서 **재무상태표·손익계산서·자본변동표·현금흐름표**
네 가지를 찾아 표 그대로 엑셀 한 파일에 담는다. 재무제표별로 시트를 나누고,
연결·별도가 모두 들어 있으면 `재무상태표(연결)` `재무상태표(별도)` 로 각각 만든다.

[왜 원문을 읽는가]
  재무제표 API(fnlttSinglAcntAll)는 계정과 금액만 준다. 표의 생김새 — 계정
  들여쓰기, 주석번호, 자본변동표의 다단 머리글 — 는 원문에만 있다. 조서에
  붙이려면 원문 표 그대로가 필요하다. 또 사업보고서를 내지 않는 비상장
  외감법인은 API에 자료가 아예 없어 감사보고서 원문이 유일한 출처다.

[표를 어떻게 찾는가 — 세 단계로 좁힌다]
  ① 제목 : '연결 재무상태표' 같은 글자를 문서에서 찾는다. 사업보고서는
           <TITLE> 태그에, 감사보고서는 본문 글자에 들어 있고 '연 결 재 무 상 태 표'
           처럼 글자 사이가 벌어져 있기도 해서 공백을 허용해 찾는다.
  ② 위치 : 표마다 '바로 앞의 제목'을 붙인다. 표를 먼저 잡고 제목을 찾는 것이
           아니라 그 반대다 — 목차에도 같은 제목이 있어서, 제목 뒤의 첫 표를
           집으면 목차 뒤에 오는 엉뚱한 표를 잡는다.
  ③ 내용 : 그 재무제표라면 반드시 있어야 하는 계정이 첫 칸에 있는지 본다
           (재무상태표면 자산총계·부채총계·자본총계). ②만으로는 감사의견
           본문의 표가 목차 제목에 딸려 들어오는 것을 막지 못한다.
  세 가지를 모두 통과한 것 중 **문서에서 가장 먼저 나온 표**를 쓴다.
  주석에 같은 이름의 표가 다시 나오지만(사업보고서 '27. 현금흐름표'),
  재무제표 본문이 언제나 주석보다 앞에 있다.

[미국(EDGAR)은 대상이 아니다]
  10-K 원문은 영문 서식이고 재무제표 이름·구성이 달라 여기서 다루지 않는다.
"""

import io
import os
import re
import zipfile

import dart_viewer


TABLE_TAG = re.compile(r"<TABLE\b[^>]*>(.*?)</TABLE>", re.DOTALL | re.IGNORECASE)
ROW_TAG = re.compile(r"<TR\b[^>]*>(.*?)</TR>", re.DOTALL | re.IGNORECASE)
# 칸은 TD·TH 뿐 아니라 TE(사업보고서 본문 표)·TU(단위 칸)로도 적힌다
CELL_TAG = re.compile(r"<T[DHEU]\b([^>]*)>(.*?)</T[DHEU]>", re.DOTALL | re.IGNORECASE)
SPAN_ATTR = re.compile(r'(COLSPAN|ROWSPAN)\s*=\s*"?(\d+)"?', re.IGNORECASE)

# 태그 사이의 글자 한 덩어리 — 제목을 찾을 단위
TEXT_NODE = re.compile(r">([^<>]+)<")

UNIT_MARK = re.compile(r"단위\s*[:：]?\s*[^)\]<]{0,8}?(억원|백만원|천원|원)")


def _spaced(word):
    """'재무상태표' → '재\\s*무\\s*상\\s*태\\s*표' (감사보고서의 '재 무 상 태 표' 대응)"""
    return r"\s*".join(re.escape(ch) for ch in word)


def _heading_re(body):
    """'(연결|별도)? 이름 (연결|별도)?' 한 줄 전체와 맞는지 보는 정규식."""
    return re.compile(
        r"^[\s\dIVXⅠ-Ⅹ.\-()]*"                    # '2-1.' '27.' 'Ⅲ.' 같은 번호
        r"(?P<pre>연\s*결|별\s*도)?\s*"
        rf"(?:{body})"
        r"\s*(?:\(\s*(?P<post>연\s*결|별\s*도)\s*(?:기\s*준)?\s*\))?"   # '현금흐름표 (연결)'
        r"[\s.\-–—…]*$"
    )


def _any(*names):
    return "|".join(_spaced(n) for n in names)


# 찾을 재무제표 — (키, 시트이름, 제목 정규식, 표에 반드시 있어야 하는 계정)
#
# ★ 손익계산서와 포괄손익계산서를 한 시트에 둔 이유
#   두 장을 따로 내는 회사(삼성전자)와 포괄손익계산서 한 장으로 끝내는
#   회사(피케이밸브)가 있다. 제목의 '포괄' 유무를 따로 기억해 두었다가
#   둘 다 있으면 같은 시트에 위아래로 이어 붙인다.
STATEMENTS = [
    ("BS", "재무상태표",
     _heading_re(_any("재무상태표", "대차대조표")),
     ("자산총계", "부채총계", "자본총계", "자산합계", "부채합계", "자본합계")),
    ("IS", "손익계산서",
     _heading_re(rf"(?P<full>{_spaced('포괄')}\s*)?{_spaced('손익계산서')}"),
     ("매출액", "영업수익", "수익(매출액)", "영업이익", "당기순이익", "총포괄손익",
      "영업이익(손실)", "당기순이익(손실)", "분기순이익", "반기순이익")),
    ("SCE", "자본변동표",
     _heading_re(_any("자본변동표")),
     ("자본금", "자본잉여금", "이익잉여금", "결손금", "주식발행초과금",
      "기타자본항목", "기초자본", "기말자본", "자본총계", "총계")),
    ("CF", "현금흐름표",
     _heading_re(_any("현금흐름표")),
     ("영업활동현금흐름", "영업활동으로인한현금흐름", "투자활동현금흐름",
      "재무활동현금흐름", "기초현금및현금성자산", "기말현금및현금성자산")),
]

# 재무제표 한 장이라면 이 정도 줄은 된다 (표지·목차·머리글 표를 걸러 낸다)
MIN_ROWS = 6


def _plain(chunk):
    """칸 안의 태그·특수공백을 걷어 낸 글자. 계정명 안의 공백은 한 칸으로 줄인다."""
    text = re.sub(r"<[^>]+>", " ", chunk)
    text = text.replace("\xa0", " ").replace("&nbsp;", " ")
    text = (text.replace("&amp;", "&").replace("&lt;", "<")
                .replace("&gt;", ">").replace("&quot;", '"'))
    return re.sub(r"\s+", " ", text).strip()


def _key(text):
    """
    계정명 대조용으로 다듬는다.
      '현금및현금성자산 (주4,29)'      → '현금및현금성자산'
      'Ⅰ. 매 출 액'                   → '매출액'
      'Ⅰ.영업활동으로 인한 현금흐름'   → '영업활동으로인한현금흐름'

    ★ 번호 접두어를 떼는 것이 중요하다.
      사업보고서 본문 표는 '매출액'이라고만 적지만, 감사보고서에 붙는
      재무제표는 'Ⅰ. 매 출 액' 처럼 번호를 달고 글자 사이를 벌려 적는다.
      떼지 않으면 감사보고서에서 재무상태표 말고는 하나도 찾지 못한다.
    """
    text = re.sub(r"\((?:주|\*)[^)]*\)", "", text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"^[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫ]+[.)]?", "", text)
    text = re.sub(r"^[IVXivx]{1,5}[.)]", "", text)
    text = re.sub(r"^\(?\d{1,2}\)?[.)]", "", text)
    text = re.sub(r"^[가-힣][.)]", "", text)
    text = re.sub(r"^[①-⑳]", "", text)
    return text.strip()


# 계정명 앞의 들여쓰기 — DART 원문은 전각공백(U+3000)으로 하위 항목을 들여쓴다
INDENT = re.compile(r"^[　\xa0\t ]+")


def _depth(chunk):
    """
    계정명 칸의 들여쓰기 깊이. 0이면 상위(합계) 항목이다.

    ★ 이것이 있어야 '소계와 그 내역'을 가려낼 수 있다.
      현금흐름표의 '재무활동현금흐름'과 그 아래 '단기차입금의 순증가',
      자본변동표의 '총포괄손익'과 그 아래 '당기순이익'은 겉보기에 똑같은
      한 줄이지만 더하면 이중계상이다. 합계검증이 멀쩡한 표를 '차이'로
      찍는 원인이 전부 이것이었다.

    전각공백·비분리공백·탭만 센다. 보통 공백과 줄바꿈은 원문 서식에서
    의미 없이 들어오는 경우가 있어 두 칸을 한 단계로만 본다.
    """
    text = re.sub(r"<[^>]+>", "", chunk).replace("&nbsp;", "\xa0")
    text = text.replace("\n", "").replace("\r", "")
    lead = INDENT.match(text)
    if not lead:
        return 0
    mark = lead.group(0)
    return (mark.count("　") + mark.count("\xa0") + mark.count("\t")
            + mark.count(" ") // 2)


def _grid(table_html):
    """
    표 하나를 [[칸, 칸, ...], ...] 격자로 편다.
    반환: (격자, 줄별 들여쓰기 깊이)

    ★ 병합된 칸은 펼쳐서 같은 글자를 채운다.
      자본변동표 머리글이 '지배기업의 소유주지분'(5칸 병합) / '자본금 | 주식발행
      초과금 | …' 처럼 2단이라, 병합을 펴지 않으면 머리글과 금액 열이 어긋나
      어느 금액이 어느 항목인지 알 수 없게 된다.
    """
    grid, depths, carry = [], [], {}   # carry: (행, 열) -> 위에서 내려온 글자
    for r, row in enumerate(ROW_TAG.finditer(table_html)):
        line, col, depth = [], 0, 0
        for attrs, inner in CELL_TAG.findall(row.group(1)):
            while (r, col) in carry:
                line.append(carry.pop((r, col)))
                col += 1

            spans = {k.upper(): int(v) for k, v in SPAN_ATTR.findall(attrs)}
            colspan = max(1, min(spans.get("COLSPAN", 1), 40))
            rowspan = max(1, min(spans.get("ROWSPAN", 1), 40))
            text = _plain(inner)
            if not line:
                depth = _depth(inner)

            for _ in range(colspan):
                line.append(text)
                for down in range(1, rowspan):
                    carry[(r + down, col)] = text
                col += 1

        while (r, col) in carry:
            line.append(carry.pop((r, col)))
            col += 1

        if line:
            grid.append(line)
            depths.append(depth)
    return grid, depths


def _looks_like(grid, keywords):
    """
    그 재무제표라면 반드시 있는 계정이 표에 있는지.

    첫 칸(계정명)뿐 아니라 위쪽 머리글 줄도 함께 본다 — 자본변동표는
    '자본금·자본잉여금·이익잉여금'이 첫 칸이 아니라 머리글에 가로로 놓인다
    (리벨리온 감사보고서). 첫 칸만 보면 자본변동표를 통째로 놓친다.
    """
    names = {_key(line[0]) for line in grid if line and line[0]}
    for line in grid[:4]:
        names.update(_key(cell) for cell in line if cell)
    return any(k in names for k in keywords)


def _headings(text):
    """
    [(위치, 키, 연결여부, 갈래)] — 문서에 나온 재무제표 제목을 순서대로.

    갈래는 손익계산서에서만 쓴다 — '포괄' 이 붙은 장과 붙지 않은 장을 구분해
    둘 다 남기기 위함이다.
    """
    found = []
    for node in TEXT_NODE.finditer(text):
        chunk = node.group(1)
        if not chunk.strip() or len(chunk) > 60:
            continue
        for key, _label, pattern, _kw in STATEMENTS:
            m = pattern.match(chunk.strip())
            if not m:
                continue
            mark = (m.group("pre") or m.group("post") or "").replace(" ", "")
            variant = "포괄" if ("full" in m.groupdict() and m.group("full")) else ""
            found.append((node.start(), key, mark == "연결", variant))
            break
    return found


def _unit(text, position):
    """표 앞쪽에서 가장 가까운 '(단위 : 백만원)' 표기."""
    marks = [m.group(0) for m in UNIT_MARK.finditer(text, 0, position)]
    return re.sub(r"\s+", " ", marks[-1]).strip() if marks else ""


def _pick(text, found):
    """
    문서 하나에서 재무제표 표를 골라 found 에 채운다.

    found : {(키, 연결여부): {갈래: {...}}}
      같은 갈래가 이미 담겼으면 건드리지 않는다 — 먼저 나온 것이 재무제표 본문이고
      뒤에 다시 나오는 같은 이름의 표는 주석이다(사업보고서 '27. 현금흐름표').
    """
    headings = _headings(text)
    if not headings:
        return

    starts = [h[0] for h in headings]
    keywords = {key: kw for key, _label, _pat, kw in STATEMENTS}

    for table in TABLE_TAG.finditer(text):
        # 이 표 바로 앞의 제목을 찾는다 (제목 뒤의 첫 표를 잡으면 목차에 걸린다)
        index = _last_before(starts, table.start())
        if index is None:
            continue
        position, key, consolidated, variant = headings[index]

        slot = found.setdefault((key, consolidated), {})
        if variant in slot:
            continue

        # ★ 포괄손익계산서는 손익계산서 '바로 다음 제목'일 때만 받는다.
        #   재무제표 본문에서는 두 장이 잇달아 나오지만, 주석에도 '15. 포괄손익
        #   계산서' 같은 표가 다시 나온다(리벨리온 감사보고서). 사이에 자본변동표·
        #   현금흐름표 제목이 끼어 있으면 본문이 아니라 주석이다.
        if variant == "포괄" and "" in slot and index != slot[""]["heading"] + 1:
            continue

        grid, depths = _grid(table.group(1))
        if len(grid) < MIN_ROWS or max((len(r) for r in grid), default=0) < 2:
            continue
        if not _looks_like(grid, keywords[key]):
            continue

        slot[variant] = {"grid": grid, "depths": depths,
                         "unit": _unit(text, table.start()),
                         "title": _title_at(text, position), "heading": index}


def _last_before(starts, position):
    """starts(오름차순)에서 position 보다 작은 마지막 값의 색인."""
    low, high = 0, len(starts)
    while low < high:
        mid = (low + high) // 2
        if starts[mid] < position:
            low = mid + 1
        else:
            high = mid
    return low - 1 if low else None


def _title_at(text, position):
    node = TEXT_NODE.match(text, position)
    return re.sub(r"\s+", " ", node.group(1)).strip() if node else ""


def extract(zip_bytes):
    """
    공시원문 zip → {(키, 연결여부): {갈래: {"grid", "unit", "title"}}}

    zip 안에 본문과 첨부(감사보고서)가 함께 들어 있는 경우가 있어 이름순으로
    돈다. 본문 파일명이 '접수번호.xml'로 첨부('접수번호_00760.xml')보다 앞서므로
    본문에서 찾은 것이 먼저 담긴다.
    """
    found = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for name in sorted(n for n in z.namelist() if n.lower().endswith(".xml")):
            text, _enc = dart_viewer.decode(z.read(name))
            _pick(text, found)
    # 제목만 있고 표를 못 찾은 자리는 지운다(_pick 이 만들어 둔 빈 칸)
    return {slot: tables for slot, tables in found.items() if tables}


def _sheets(found):
    """[(시트이름, [표, ...])] — 재무제표 순서 · 연결 먼저 · 손익은 포괄을 뒤에."""
    out = []
    for key, label, _pattern, _kw in STATEMENTS:
        for consolidated in (True, False):
            tables = found.get((key, consolidated))
            if not tables:
                continue
            ordered = [tables[v] for v in ("", "포괄") if v in tables]
            out.append((f"{label}({'연결' if consolidated else '별도'})", ordered))
    return out


# ─────────────────────────────────────────────────────────────
# 엑셀로 쓰기
# ─────────────────────────────────────────────────────────────

NUMBER = re.compile(r"^\(?-?[\d,]+(\.\d+)?\)?$")


def _cell_value(text):
    """'1,234' → 1234, '(1,234)' → -1234. 숫자가 아니면 글자 그대로."""
    if not text or not NUMBER.match(text):
        return text
    negative = text.startswith("(") and text.endswith(")")
    body = text.strip("()").replace(",", "")
    try:
        value = float(body)
    except ValueError:
        return text
    if value.is_integer():
        value = int(value)
    return -value if negative else value


# ─────────────────────────────────────────────────────────────
# 합계검증 — 재무제표가 스스로 맞는지
# ─────────────────────────────────────────────────────────────
# 표를 글자로 읽는 방식이라 '한 줄을 빠뜨리지 않았는가'를 눈으로 확인할 길이
# 필요하다. 재무제표는 스스로 맞아야 하는 항등식을 갖고 있으므로 그것을 그대로
# 검산해 「합계검증」 시트에 적는다. 파싱이 잘못되면 여기서 '차이'로 드러난다.
#
# ★ 조서에 옮기기 전 대조용이기도 하다 — 회사가 낸 표 자체가 어긋나는 경우
#   (표시 반올림, 중단영업 별도 표시 등)도 여기서 보인다.

# 항등식 판정 허용오차 — 표시 단위 반올림만 봐준다
TOLERANCE = 1.0


def _num(values, *names):
    """계정명 후보 중 먼저 잡히는 값. 없으면 None."""
    for name in names:
        if name in values:
            return values[name]
    return None


def _find(values, pattern):
    """정규식으로 계정을 찾는다(회사마다 이름이 조금씩 달라서)."""
    for name, value in values.items():
        if re.search(pattern, name):
            return value
    return None


def _columns(grid):
    """
    열마다 {계정키: 금액}.

    ★ 감사보고서 재무제표는 한 기수를 두 열로 적는다(세부 금액 열 + 소계 열).
      그래서 '열 하나 = 한 기수'가 아니라 열마다 따로 모은다. 검산에 쓰는
      항목(자산총계·영업활동현금흐름 등)은 모두 같은 소계 열에 모이므로
      열 단위로 검산하면 그대로 맞아떨어진다.
    """
    columns = {}
    for line in grid:
        if not line or not line[0]:
            continue
        key = _key(line[0])
        for index, text in enumerate(line[1:], start=1):
            value = _cell_value(text)
            if isinstance(value, (int, float)):
                columns.setdefault(index, {}).setdefault(key, value)
    return columns


def _column_labels(grid):
    """열 이름('제 57 기') — 금액이 나오기 전의 머리글 줄에서 가져온다."""
    labels = {}
    for line in grid:
        if any(isinstance(_cell_value(t), (int, float)) for t in line[1:]):
            break
        for index, text in enumerate(line[1:], start=1):
            if text:
                labels[index] = text
    return labels


def _check_bs(values):
    assets = _num(values, "자산총계", "자산합계", "자산총액")
    debts = _num(values, "부채총계", "부채합계", "부채총액")
    equity = _num(values, "자본총계", "자본합계", "자본총액")
    current = _num(values, "유동자산")
    noncurrent = _num(values, "비유동자산")
    current_debt = _num(values, "유동부채")
    noncurrent_debt = _num(values, "비유동부채")
    both = _num(values, "부채와자본총계", "부채및자본총계", "부채와자본총액")

    out = []
    if None not in (assets, debts, equity):
        out.append(("자산 = 부채 + 자본", "부채총계 + 자본총계", debts + equity, assets))
    if None not in (assets, current, noncurrent):
        out.append(("자산총계 구성", "유동자산 + 비유동자산", current + noncurrent, assets))
    if None not in (debts, current_debt, noncurrent_debt):
        out.append(("부채총계 구성", "유동부채 + 비유동부채",
                    current_debt + noncurrent_debt, debts))
    if None not in (both, assets):
        out.append(("부채와자본총계 = 자산총계", "자산총계", assets, both))
    return out


def _check_is(values):
    revenue = _num(values, "매출액", "영업수익", "수익(매출액)", "매출및지분법손익", "매출")
    cost = _num(values, "매출원가")
    gross = _num(values, "매출총이익", "매출총이익(손실)", "매출총손익")
    sga = _num(values, "판매비와관리비", "판매비및관리비", "판매관리비", "판매비와일반관리비")
    operating = _num(values, "영업이익", "영업이익(손실)", "영업손익", "영업손실")
    pretax = _find(values, r"법인세.*차감전")
    tax = _num(values, "법인세비용", "법인세비용(수익)", "법인세수익", "법인세등")
    net = _find(values, r"^(당기|분기|반기|연결|당기연결)?(연결)?순(이익|손실|손익)")
    # ★ 여기는 정확히 일치하는 이름만 본다.
    #   '기타포괄손익-공정가치측정유가증권 평가손익'처럼 앞글자가 같은 세부 계정이
    #   많아 앞부분만 맞춰 찾으면 그 줄을 총계로 잘못 집는다(신한지주).
    other = _num(values, "기타포괄손익", "기타포괄이익", "세후기타포괄손익",
                 "연결기타포괄손익", "기타포괄손익(세후)", "당기기타포괄손익")
    total = _num(values, "총포괄손익", "총포괄이익", "당기총포괄손익",
                 "연결총포괄손익", "총포괄손익(세후)", "포괄손익합계")

    out = []
    if None not in (gross, revenue, cost):
        out.append(("매출총이익", "매출액 − 매출원가", revenue - cost, gross))
    if None not in (operating, gross, sga):
        out.append(("영업이익", "매출총이익 − 판매비와관리비", gross - sga, operating))
    if None not in (net, pretax, tax):
        out.append(("당기순이익", "법인세차감전순이익 − 법인세비용", pretax - tax, net))
    if None not in (total, net, other):
        out.append(("총포괄손익", "당기순이익 + 기타포괄손익", net + other, total))
    return out


def _check_cf(values, sums=(None, None)):
    """
    현금흐름표 — 활동별 합계와 기초·기말이 맞는지.

    ★ 활동별 합계는 '증감 줄 위의 상위 단계 줄'을 모두 더해서 본다.
      영업·투자·재무 셋만 더하면 회사마다 다른 조정 줄에 걸려 멀쩡한 표가
      '차이'로 찍힌다 — 환율효과를 증감에 넣는 회사와 기말에 넣는 회사가
      갈리고(삼성전자는 전자), '매각예정분류'(FY2023 △14,153백만) 같은
      줄이 따로 서기도 한다. 상위 항목을 통째로 더하면 어느 쪽이든 맞는다.

    ★ '증감' 계정을 찾을 때 반드시 줄 첫머리부터 본다.
      '외화환산으로 인한 현금의 변동'도 '…현금의 변동'이라 가운데만 보면 걸린다.
    """
    change = _find(values, r"^현금(및현금성자산)?의(순)?(증가|감소|증감|변동)")
    opening = _find(values, r"^(기초|기초의|당기초|전기초).*현금|^현금.*기초")
    closing = _find(values, r"^(기말|기말의|당기말|전기말).*현금|^현금.*기말")

    before, after = sums

    out = []
    if before is not None and change is not None:
        # 증감 줄 위의 '상위 항목'을 전부 더한다 — 영업·투자·재무활동에
        # 환율효과·매각예정분류 같은 조정 줄까지 자동으로 들어온다
        out.append(("현금의 증감", "증감 줄 위의 상위 항목 합계", before, change))
    if None not in (closing, opening, change):
        # 환율효과를 증감이 아니라 기말 쪽에 붙이는 회사가 있어 그 줄도 더한다
        extra = after or 0.0
        out.append(("기말 현금",
                    "기초현금 + 현금의 증감" + (" + 증감 아래 조정" if extra else ""),
                    opening + change + extra, closing))
    return out


BLOCK_OPEN = re.compile(r"(기초|전기초|당기초)")
BLOCK_CLOSE = re.compile(r"(기말|전기말|당기말)")
CASH_EDGE = re.compile(r"^(기초|기말|기초의|기말의|당기초|당기말|전기초|전기말)|현금.*(기초|기말)")


def _rows_at(grid, depths, column):
    """[(계정키, 깊이, 그 열의 값)] — 검산은 이 셋만 있으면 된다."""
    out = []
    for line, depth in zip(grid, depths):
        name = _key(line[0]) if line and line[0] else ""
        value = _cell_value(line[column]) if column < len(line) else None
        out.append((name, depth, value if isinstance(value, (int, float)) else None))
    return out


CHANGE_ROW = re.compile(r"^현금(및현금성자산)?의(순)?(증가|감소|증감|변동)")


SUBTOTAL_NAME = re.compile(r"(합\s*계|소\s*계)$|^총|^계$")


def _top_level(rows, span):
    """
    span 구간에서 '가장 얕은 들여쓰기' 단계를 찾는다.

    0단으로 고정하면 안 된다 — 표 전체를 한 단 들여쓰는 회사가 있어
    (신한지주 자본변동표) 그러면 더할 줄이 하나도 없게 된다.
    """
    levels = [d for _n, d, v in rows[span[0]:span[1]] if v is not None]
    return min(levels) if levels else None


def _sum_without_double(rows, exclude=None):
    """
    소계와 그 내역이 섞인 줄들을 이중계상 없이 더한다.

    ★ 두 가지 모양을 모두 받아야 한다.
        (가) 소계가 위에 있고 내역이 한 단 더 들어감 (비에이치아이 '총포괄손익')
             → 더 깊은 줄을 건너뛰면 된다.
        (나) 소계와 내역이 같은 깊이에 나란히 있음 (신한지주 '총포괄이익' 아래
             같은 깊이의 '당기순이익·기타포괄손익')
             → 이름이 '…합계·총…' 이면 같은 깊이의 형제까지 건너뛴다.

    빈 칸(값이 없는 줄)은 구역 머리글이라 건너뛰되, 그 아래를 다시 열어 준다.
    """
    total, deeper, siblings = 0.0, None, None
    for index, (name, depth, value) in enumerate(rows):
        if siblings is not None and depth < siblings:
            siblings = None
        if siblings is not None and depth == siblings:
            continue
        if deeper is not None and depth > deeper:
            continue
        if value is None:
            deeper = None
            continue
        if exclude and exclude.match(name):
            continue
        total += value
        deeper = depth
        # 소계의 내역이 같은 깊이에 나란히 오는 표에서만 형제를 건너뛴다.
        # 내역을 한 단 들여쓰는 표(비에이치아이)에서 형제까지 건너뛰면
        # '지배력을 상실하지 않는 종속기업…' 같은 별개 항목이 빠진다.
        nested = index + 1 < len(rows) and rows[index + 1][1] > depth
        if SUBTOTAL_NAME.search(name) and not nested:
            siblings = depth
    return total


def _cash_sums(grid, depths, column):
    """
    현금흐름표의 상위 항목 합계를 (증감 줄 위, 증감~기말 사이) 로 나눠 준다.

      증감 줄 위   : 영업·투자·재무활동 + 회사별 조정 줄
      증감 줄 아래 : 환율효과를 기말 쪽에 붙이는 회사의 그 줄

    들여쓰기가 전혀 없는 표는 소계와 내역을 가릴 수 없어 (None, None).
    """
    rows = _rows_at(grid, depths, column)
    if not any(depth for _n, depth, _v in rows):
        return None, None

    change_at = closing_at = None
    for index, (name, _depth, value) in enumerate(rows):
        if not name or value is None:
            continue
        if change_at is None and CHANGE_ROW.match(name):
            change_at = index
        elif change_at is not None and re.match(r"^(기말|기말의|당기말|전기말)|^현금.*기말", name):
            closing_at = index
            break
    if change_at is None:
        return None, None

    def add(span):
        level = _top_level(rows, span)
        if level is None:
            return 0.0
        return sum(v for n, d, v in rows[span[0]:span[1]]
                   if v is not None and d == level and not CASH_EDGE.match(n))

    before = add((0, change_at))
    after = add((change_at + 1, closing_at)) if closing_at else 0.0
    return before, after


def _check_sce(grid, depths, column):
    """
    자본변동표 — 기말자본 = 기초자본 + 기중 변동 합계.

    한 표에 전기 블록과 당기 블록이 위아래로 들어 있으므로 '기초' 줄부터
    '기말' 줄까지를 한 덩어리로 보고 각각 검산한다.

    ★ 들여쓰기 0단인 줄만 더한다. '총포괄손익'(소계) 아래에 '당기순이익·
      자산재평가손익'이 딸려 오는 표가 흔해서(비에이치아이·신한지주),
      전부 더하면 소계만큼 이중계상된다.
    """
    rows = _rows_at(grid, depths, column)

    out, start = [], None
    for index, (name, _depth, _value) in enumerate(rows):
        if not name:
            continue
        if BLOCK_OPEN.search(name):
            start = index
        elif BLOCK_CLOSE.search(name) and start is not None and index > start:
            opening, closing = rows[start][2], rows[index][2]
            middle = _sum_without_double(rows[start + 1:index])
            if opening is not None and closing is not None:
                out.append((f"{grid[index][0]} 자본", "기초자본 + 기중 변동 합계",
                            opening + middle, closing))
            start = None
    return out


def verify(found):
    """
    [(시트, 기수, 항목, 산식, 계산값, 표시값, 차이, 판정)] 을 만든다.

    검산할 항목이 표에 없으면(서비스업의 매출원가 등) 조용히 건너뛴다 —
    없는 것과 틀린 것은 다르므로 '차이'로 적지 않는다.
    """
    rows = []
    for name, tables in _sheets(found):
        key = "BS" if name.startswith("재무상태표") else \
              "IS" if name.startswith("손익계산서") else \
              "SCE" if name.startswith("자본변동표") else "CF"

        for item in tables:
            grid = item["grid"]
            depths = item.get("depths") or [0] * len(grid)
            labels = _column_labels(grid)
            for column, values in sorted(_columns(grid).items()):
                if key == "BS":
                    checks = _check_bs(values)
                elif key == "IS":
                    checks = _check_is(values)
                elif key == "CF":
                    checks = _check_cf(values, _cash_sums(grid, depths, column))
                else:
                    checks = _check_sce(grid, depths, column)

                for label, formula, computed, shown in checks:
                    gap = computed - shown
                    rows.append((name, labels.get(column, f"{column}열"),
                                 label, formula, computed, shown, gap,
                                 "일치" if abs(gap) <= TOLERANCE else "차이"))
    return rows


VERIFY_HEAD = ["재무제표", "기수·열", "검증항목", "산식", "계산값", "표시값", "차이", "판정"]


def _write_verification(book, rows, meta, head_font, head_fill, label_font):
    """맨 앞에 「합계검증」 시트를 만든다. 검산할 게 없으면 만들지 않는다."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if not rows:
        return

    sheet = book.create_sheet("합계검증")
    line = 1
    for key, value in (meta or {}).items():
        sheet.cell(row=line, column=1, value=key).font = label_font
        sheet.cell(row=line, column=2, value=value)
        line += 1

    gaps = [r for r in rows if r[7] == "차이"]
    sheet.cell(row=line, column=1, value="검증 결과").font = label_font
    sheet.cell(row=line, column=2,
               value=f"{len(rows)}건 중 {len(rows) - len(gaps)}건 일치"
                     + (f" · {len(gaps)}건 차이 — 아래 빨간 줄을 확인하십시오"
                        if gaps else " · 전건 일치"))
    line += 2

    sheet.cell(row=line, column=1,
               value="재무제표가 스스로 만족해야 하는 항등식을 원문 표의 숫자로 검산한 것입니다. "
                     "'차이'는 표를 잘못 읽었거나 회사 표시가 어긋난다는 뜻이므로 원문을 확인하십시오.")
    line += 2

    header = line
    for column, name in enumerate(VERIFY_HEAD, start=1):
        cell = sheet.cell(row=header, column=column, value=name)
        cell.font = head_font
        cell.fill = head_fill
    line += 1

    bad_font = Font(bold=True, color="C0392B")
    for record in rows:
        for column, value in enumerate(record, start=1):
            cell = sheet.cell(row=line, column=column, value=value)
            if record[7] == "차이":
                cell.font = bad_font
        line += 1

    for column, width in enumerate((20, 22, 22, 30, 18, 18, 12, 8), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = sheet.cell(row=header + 1, column=1)
    sheet.auto_filter.ref = f"A{header}:H{line - 1}"


def write(found, path, meta=None):
    """
    골라 낸 재무제표를 엑셀 한 파일로 쓴다. 시트가 하나도 없으면 만들지 않고
    None을 돌려준다. 맨 앞에 「합계검증」 시트를 붙인다.

    meta : 시트 맨 위에 적을 출처 정보 {"회사": ..., "보고서": ..., ...}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    sheets = _sheets(found)
    if not sheets:
        return None

    book = Workbook()
    book.remove(book.active)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="00338D")
    label_font = Font(bold=True)
    wrap = Alignment(vertical="center", wrap_text=True)

    _write_verification(book, verify(found), meta,
                        head_font, head_fill, label_font)

    for name, tables in sheets:
        sheet = book.create_sheet(name[:31])
        row = 1

        for key, value in (meta or {}).items():
            sheet.cell(row=row, column=1, value=key).font = label_font
            sheet.cell(row=row, column=2, value=value)
            row += 1

        widest, freeze_at = 1, None
        for item in tables:                 # 손익계산서 + 포괄손익계산서
            row += 1                        # 앞 표와 한 줄 띄우기
            if item["title"]:
                sheet.cell(row=row, column=1, value=item["title"]).font = label_font
                if item["unit"]:
                    sheet.cell(row=row, column=2, value=f"({item['unit']})")
                row += 1

            first = row
            for line in item["grid"]:
                for col, text in enumerate(line, start=1):
                    # 첫 칸(계정명)은 '1,234'처럼 보여도 글자 그대로 둔다
                    sheet.cell(row=row, column=col,
                               value=text if col == 1 else _cell_value(text))
                row += 1

            # 표의 머리글 줄에 색을 넣는다 — 금액이 아직 안 나온 위쪽 줄이 머리글이다
            for offset, line in enumerate(item["grid"]):
                if any(isinstance(_cell_value(c), (int, float)) for c in line[1:]):
                    break
                for col in range(1, len(line) + 1):
                    cell = sheet.cell(row=first + offset, column=col)
                    cell.font = head_font
                    cell.fill = head_fill
                    cell.alignment = wrap

            widest = max(widest, max((len(l) for l in item["grid"]), default=1))
            freeze_at = freeze_at or first

        sheet.column_dimensions["A"].width = 46
        for col in range(2, widest + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = 16
        if freeze_at:
            sheet.freeze_panes = sheet.cell(row=freeze_at, column=2)

    book.save(path)
    return path


def build(zip_bytes, path, meta=None):
    """원문 zip 하나를 받아 엑셀을 만든다. 찾은 게 없으면 None."""
    return write(extract(zip_bytes), path, meta)


def summary(found):
    """로그에 적을 '재무상태표(연결) · 손익계산서(연결) …' 문자열."""
    return " · ".join(name for name, _tables in _sheets(found))


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("사용법: python fs_excel.py <공시원문.zip> [저장경로.xlsx]")
        sys.exit(1)

    with open(sys.argv[1], "rb") as f:
        blob = f.read()
    out = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.splitext(sys.argv[1])[0] + "_재무제표.xlsx"

    got = extract(blob)
    print("찾음:", summary(got) or "없음")
    if got:
        print("저장:", write(got, out))
